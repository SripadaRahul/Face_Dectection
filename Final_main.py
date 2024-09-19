# import libraries
################################### lib for TKINTER GUI
from tkinter import *
from tkinter import messagebox
import tkinter.messagebox as tmsg
import shutil,os
from tkinter.filedialog import askopenfilename
from tkinter import *
################################### importing scripts
from tkinter import filedialog
from tkinter.filedialog import askopenfilename

import detect as dt
import Dates as D
################################### lib to perform CRUD operations in csv/excel
from PIL import Image, ImageTk
from openpyxl import *
import pandas as pd
import numpy as np
import csv
################################### other depenmaildencies
import datetime
import time
import os
import sys
###################################lib to automate the mail
import smtplib
from email.message import EmailMessage
sys.path.insert(1,'C:\\Users\\sripa\\PycharmProjects\\Face\\data\\Attendance_xlsx')
from data.Attendance_xlsx import automail

check=100

'''*************************************************************** Code for GUI starts ****************************************************************'''

'''*************************************************************** initialize Window ****************************************************************'''

# creating tkinter window
root = Tk()

# creating fixed geometry of the
# tkinter window with dimensions 150x200
root.geometry("150x200")
root.maxsize(1060,550)
root.minsize(1060,550)

#providing a title
root.title("ACE Engineering College Attendance System...")

image = Image.open("GUI/register.png")
image=image.resize((1060,550), Image.ANTIALIAS)
photo = ImageTk.PhotoImage(image)

''' ************************************************************* Code for Registeration starts ********************************************************'''
def register():
    global register_screen
    register_screen = Toplevel(root)
    register_screen.title("Register")
    register_screen.geometry("300x250")

    global username
    global password
    global username_entry
    global password_entry
    username = StringVar()
    password = StringVar()

    Label(register_screen, text="Please enter details below", bg="yellow").pack()
    Label(register_screen, text="").pack()
    username_lable = Label(register_screen, text="Username * ")
    username_lable.pack()
    username_entry = Entry(register_screen, textvariable=username)
    username_entry.pack()
    password_lable = Label(register_screen, text="Password * ")
    password_lable.pack()
    password_entry = Entry(register_screen, textvariable=password, show='*')
    password_entry.pack()
    Label(register_screen, text="").pack()
    Button(register_screen, text="Register", width=10, height=1, bg="blue", command=register_user).pack()


# Designing window for login



def register_user():
    username_info = username.get()
    password_info = password.get()

    wb = load_workbook('C:\\Users\\sripa\\PycharmProjects\\Face\\teacherreg.xlsx')

    sheet = wb.active
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 15


    sheet.cell(row=1, column=1).value = "username"
    sheet.cell(row=1, column=2).value = "password"


    current_row = sheet.max_row
    current_column = sheet.max_column

    sheet.cell(row=current_row + 1, column=1).value = username.get()
    sheet.cell(row=current_row + 1, column=2).value = password.get()


    wb.save('C:\\Users\\sripa\\PycharmProjects\\Face\\teacherreg.xlsx')

    print(username.get(), " is Registered successfully")

    username_entry.delete(0, END)
    password_entry.delete(0, END)

    Label(register_screen, text="Registration Success", fg="green", font=("calibri", 11)).pack()



''' ************************************************************* Code for Registeration ENDS ********************************************************'''
''' ************************************************************* Code for Login window starts ********************************************************'''




def login_verify():
    user1=lid.get()
    pass1 = lpass.get()
    e1.delete(0, END)
    e2.delete(0, END)
    wb = load_workbook('C:\\Users\\sripa\\PycharmProjects\\Face\\teacherreg.xlsx')

    sheet = wb.active
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 15



    cola=sheet['A']
    colb=sheet['B']





    for cella in cola:
        for cellb in colb:
            if cella.value==user1 and cellb.value == pass1:
                Proceed_menu()
                return 1
    return 0
    # if err==0:
    #     messagebox.showwarning("Warning", "Wrong User_Id/Password!")







# Designing popup for login invalid password

def password_not_recognised():
    global password_not_recog_screen
    password_not_recog_screen = Toplevel(root)
    password_not_recog_screen.title("Success")
    password_not_recog_screen.geometry("150x100")
    Label(password_not_recog_screen, text="Invalid Password ").pack()
    Button(password_not_recog_screen, text="OK", command=delete_password_not_recognised).pack()


# Designing popup for user not found

def user_not_found():
    global user_not_found_screen
    user_not_found_screen = Toplevel(root)
    user_not_found_screen.title("Success")
    user_not_found_screen.geometry("150x100")
    Label(user_not_found_screen, text="User Not Found").pack()
    Button(user_not_found_screen, text="OK", command=delete_user_not_found_screen).pack()

def delete_password_not_recognised():
    password_not_recog_screen.destroy()


def delete_user_not_found_screen():
    user_not_found_screen.destroy()

global login_by
Admin_id="admin"
Admin_pass="admin"
newuser_id="ace"
newuser_pass="ace"

def Login():
    global login_by
    lid1=lid.get()
    lpass1=lpass.get()

    if(lid1==Admin_id and lpass1==Admin_pass):
        login_by="Admin"
        Proceed_menu()
        messagebox.showinfo("Information", "Logged in as Administrator")
    elif (lid1 == newuser_id and lpass1 ==newuser_pass):
        login_by = "Tester"
        Proceed_menu()
        messagebox.showinfo("Information", "Logged in as Tester")
    else:
        login_by=""
        if login_verify()==1:
            messagebox.showinfo("Information", "Logged in as User")
        else:
            messagebox.showwarning("Information", "Wrong User_Id/Password")

l_login=Label(image=photo)
f_login=Frame(l_login,pady="25",padx="25") #cretaing a Frame which can expand according to the size of the window
global lid
global lpass
lb0 =Label(f_login,text="Enter Details",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(columnspan=3,row=0,pady="15")
lb1 =Label(f_login,text="User_ID: ",font="lucida 10 bold").grid(column=0,row=2,pady="4")
global e1
global e2
lid =StringVar()
e1 =Entry(f_login,textvariable=lid,width="28")
e1.grid(column=1,row=2)
lb2 =Label(f_login,text="User_Password: ",font="lucida 10 bold").grid(column=0,row=3,pady="4")

lpass=StringVar()
e2=Entry(f_login,show="*",textvariable=lpass,width="28")
e2.grid(column=1,row=3)
btn=Button(f_login,text="login",bg="green",fg="white",width="10",font="lucida 10 bold",command=Login)
btn.grid(column=0,row=5,pady="10")
regbtn = Button(f_login, bg="green",fg="white",width="10",font="lucida 10 bold",text='Register', command=register)
regbtn.grid(columnspan=3,column =1,row=5,pady="10")
f_login.pack(pady="165")

l_login.pack(ipadx="100",fill=BOTH)

''' ************************************************************* Code for Login window Ends ********************************************************'''
''' ************************************************************* Code for Menu Widgit starts ********************************************************'''
def admin(x):
    if (x == 1):
        Uploadi()

    if (x == 2):
        deletei()

def teacher(t):
    if t==1:
        register()

def attendance(y):
    l.pack_forget()

    l1.pack(ipadx="150", fill=BOTH)
    if (y == 1):
        fa.pack_forget()
        fd.pack(pady="120")
    if (y == 2):
        fd.pack_forget()
        fa.pack(pady="135")

def more(z):
    l.pack_forget()
    l1.pack_forget()
    l2.pack(ipadx="100",fill=BOTH)




def Proceed_menu():
    global login_by
    global login_by
    l_login.pack_forget()

    mainmenu = Menu(root)
    root.config(menu=mainmenu)

    if (login_by == "Admin" or login_by=="Tester"):#-------------------Only access to admin-----------------------
        #-------------For registering new users-------------------------
        m1 = Menu(mainmenu, tearoff=0)
        m1.add_command(label="Register", command=lambda: teacher(1))
        mainmenu.add_cascade(label="Teacher", menu=m1)

    m2 = Menu(mainmenu, tearoff=0)
    m2.add_command(label="Detect", command=lambda: attendance(1))
    m2.add_separator()
    m2.add_command(label="View Attendance", command=lambda: attendance(2))
    root.config(menu=mainmenu)
    mainmenu.add_cascade(label="Attendance", menu=m2)

    if (login_by == "Admin"or login_by=="Tester"):

        #-----------for Mail Button------------------------------------
        mailbtn = Button(root, text='Send Mail', command=lambda: mail(1)).place(x=500, y=50)

    # -------------For Adding/Deleting Images-----------------------
    m4 = Menu(mainmenu, tearoff=0)

    m4.add_command(label="Upload Image", command=lambda: admin(1))
    m4.add_command(label="Delete Image", command=lambda: admin(2))
    root.config(menu=mainmenu)
    mainmenu.add_cascade(label="Admin", menu=m4)

    # -------------For HELP-----------------------
    m3 = Menu(mainmenu, tearoff=0)
    m3.add_command(label="Help", command=lambda: more(1))
    # m3.add_command(label="About Us", command=lambda: more(1))
    root.config(menu=mainmenu)
    mainmenu.add_cascade(label="More", menu=m3)







    # mailbtn = Button(root, text='Send Mail', command=lambda:mail(1)).place(x=500, y=50)



''' ************************************************************* Code for Menu Widgit Ends ********************************************************'''
''' ************************************************************* Code for Mail Button starts********************************************************'''

def mail(args):
    if(args==1):
        res = tmsg.askquestion('Type of Mail','Send Mail Manually')
        global check
        if res == 'yes':
            check=0
            automail.mail(0)

        else:
            pass
            check=1
            messagebox.showinfo("Information", "Time set for Automail is Done")



''' ************************************************************* Code for Mail Button Ends ********************************************************'''


''' ************************************************************* Code for Upload image starts ********************************************************'''


def Uploadi():
    try:
        src = askopenfilename()

        des = 'C:\\Users\\sripa\\PycharmProjects\\Face\\ImagesAttendance'
        shutil.copy(src, des)
        print('SUCCESSFULLY UPLOADED in the destination folder now you can detect new students!!!')
    except:
        print("No Image Uploaded")

errmsg = 'Error!'

''' ************************************************************* Code for Upload images Ends ********************************************************'''
''' ************************************************************* Code for Delete image starts ********************************************************'''
def deletei():
    try:
        filename = filedialog.askopenfilename(initialdir='C:\\Users\\sripa\\PycharmProjects\\Face\\ImagesAttendance',
                                              title="Select Image")
        os.remove(filename)
        print("Image Deleted Successfully!!!")
    except:
        print("No Image Deleted")

''' ************************************************************* Code for Delete images Ends ********************************************************'''
''' ************************************************************* Code for Recording attendance starts ********************************************************'''


def detect1():
    global check
    check =2

    insertdate(desem.get()[0], desection.get())
    dt.detect(desem.get()[0], desection.get())

l = Label(image=photo)
l.pack(ipadx="100", fill=BOTH)

l1 = Label(image=photo)

fd = Frame(l1, pady="25", padx="25")
ld = Label(fd, text="This is Detect Section", bg="orange", fg="blue", font="lucida 10 bold", width="35", pady="4").grid(
    columnspan=3, row=0, pady="15")
l4 = Label(fd, text="Sem", font="lucida 10 bold").grid(column=0, row=1, pady="4")

desem = StringVar()
desem.set("1st sem")  # default value

w1 = OptionMenu(fd, desem, "1st sem", "2nd sem", "3rd sem", "4th sem", "5th sem", "6th sem", "7th sem", "8th sem").grid(
    column=1, row=1, pady="4")
l5 = Label(fd, text="Section", font="lucida 10 bold").grid(column=0, row=2, pady="4")

desection = StringVar()
w2 = OptionMenu(fd, desection, "A", "B").grid(column=1, row=2, pady="4")

b1 = Button(fd, text="Detect", bg="green", fg="white", width="10", font="lucida 10 bold", command=detect1)
b1.grid(columnspan=3, row=3, pady="20")

fd.pack(pady="120")


def insertdate(sem, sec):
    if sem == '1' or sem == '2':
        year = 'first_year'
    elif sem == '3' or sem == '4':
        year = 'second_year'
    elif sem == '5' or sem == '6':
        year = 'third_year'
    else:
        year = 'fourth_year'
    flag = 0
    print('Checking if the date is working or not..')
    for i in D.filterdates():
        if str(i.day) == str(datetime.datetime.today().day) and str(i.month) == str(
                datetime.datetime.today().month) and str(i.year) == str(datetime.datetime.today().year):
            flag = 1
    if flag == 0:
        value = tmsg.askquestion("this is a holiday.... want to continue.")
        if value == "yes":
            wb = load_workbook(f'data/Attendance_xlsx/{year}_{sem}sem_ECE_{sec}.xlsx')
            if flag == 0:
                print('Date:', str(i)[:11], ' is written in excel and is a holiday')
            else:
                print('Date:', str(i)[:11], ' is written in excel and is a working day')

            sheet = wb.active
            current_row = sheet.max_row
            current_column = sheet.max_column
            print(current_column)
            sheet.column_dimensions['A'].width = 20
            sheet.column_dimensions['B'].width = 20
            sheet.cell(row=1, column=1).value = "Name"
            sheet.cell(row=1, column=2).value = "Enrollment"

            current_row = sheet.max_row
            current_column = sheet.max_column
            # sheet.cell(row=1,column=current_column).width = 20
            sheet.cell(row=1, column=current_column + 1).value = "".join(str(datetime.datetime.today())[:11])

            # save the file
            wb.save('data/Attendance_xlsx/third_year_5sem_IT2.xlsx')


''' ************************************************************* Code for Recording attendance Ends ********************************************************'''

''' ************************************************************* Code for viewing Excel starts ********************************************************'''


def open_excel():
    sem = seme.get()[0]
    sec = sect.get()
    if sec=="1":
        sec="A"
    elif sec=="2":
        sec="B"
    if sem == '1' or sem == '2':
        year = 'first_year'
    elif sem == '3' or sem == '4':
        year = 'second_year'
    elif sem == '5' or sem == '6':
        year = 'third_year'
    else:
        year = 'fourth_year'
    file = "C:\\Users\\sripa\\PycharmProjects\\Face\\data\\Attendance_xlsx\\" + year + "_" + sem + "sem_ECE_" + sec + ".xlsx"
    os.startfile(file)


fa = Frame(l1, pady="8", padx="20", height=200)
Label(fa, text="Select Year", bg="orange", fg="blue", font="lucida 10 bold", width="30").grid(columnspan=3, row=0,
                                                                                              pady="10")
year = StringVar()
year.set("first_year")
w1 = OptionMenu(fa, year, "first_year", "second_year", "third_year", "fourth_year").grid(columnspan=3, row=1, pady="4")

Label(fa, text="Select Semester", bg="orange", fg="blue", font="lucida 10 bold", width="30").grid(columnspan=3, row=2,
                                                                                                  pady="10")
seme = StringVar()
seme.set("1sem")  # default value
w1 = OptionMenu(fa, seme, "1sem", "2sem", "3sem", "4sem", "5sem", "6sem", "7sem", "8sem").grid(columnspan=3, row=3,
                                                                                               pady="4")

sect = StringVar()
sect.set("nosec")
Label(fa, text="Select Section", bg="orange", fg="blue", font="lucida 10 bold", width="30").grid(columnspan=3, row=5,
                                                                                                 pady="10")
radio = Radiobutton(fa, text="ECE-A", variable=sect, value="1").grid(column=0, row=6, pady="4")
radio = Radiobutton(fa, text="ECE-B", padx=14, variable=sect, value="2").grid(column=1, row=6, pady="4")
btn = Button(fa, text="show", bg="green", fg="white", width="10", font="lucida 10 bold", command=open_excel)
btn.grid(columnspan=3, row=7, pady="0")
fa.pack(pady="135")
l1.pack(ipadx="150", fill=BOTH)


''' ************************************************************* Code for viewing Excel Ends ********************************************************'''
''' ************************************************************* Code for more details starts ********************************************************'''

l2=Label(image=photo)

f=Frame(l2,pady="25",padx="25")
lbl=Label(f,text="Any Queries",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(column=0,row=0)
lbl=Label(f,text="you can contact us on following",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(column=0,row=1)
lbl=Label(f,text="Email  :  sripadarahul28@gmail.com",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(column=0,row=2)
lbl=Label(f,text="mobile : +918500428806",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(column=0,row=3)
f.pack(pady="195")

l2.pack(ipadx="100",fill=BOTH)

''' ************************************************************* Code for more details Ends ********************************************************'''




Button(root, text="Quit", command=root.destroy).pack()
root.mainloop()
# print("check "+str(check))

if check==0:
    pass
elif check==2:
    automail.mail(1)
elif check==1:
    automail.mail(1)

