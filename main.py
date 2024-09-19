# import libraries
################################### lib for TKINTER GUI
from tkinter import *
import tkinter.messagebox as tmsg
################################### importing scripts
import detect as dt
import Dates as D
################################### lib to perform CRUD operations in csv/excel
from PIL import Image, ImageTk
from openpyxl import *
import pandas as pd
import numpy as np
import csv
################################### other dependencies
import datetime
import time
import os


'''*************************************************************** Code for GUI starts ****************************************************************'''

'''*************************************************************** initialize Window ****************************************************************'''

# creating tkinter window
root = Tk()

# creating fixed geometry of the
# tkinter window with dimensions 150x200
root.geometry("150x200")
root.maxsize(580,550)
root.minsize(580,550)

#providing a title
root.title("Auto Attendance..")

image = Image.open("GUI/register.jpg")
image=image.resize((580,600), Image.ANTIALIAS)
photo = ImageTk.PhotoImage(image)


''' ************************************************************* Code for Login window starts ********************************************************'''
global login_by
Admin_id="admin"
Admin_pass="admin@123"

def Login():
    global login_by
    lid1=lid.get()
    lpass1=lpass.get()
    print(lid.get(),lpass.get())
    if(lid1==Admin_id and lpass1==Admin_pass):
        login_by="Admin"
        Proceed_menu()
    else:
        login_by=""
        Proceed_menu()

l_login=Label(image=photo)
f_login=Frame(l_login,pady="25",padx="25") #cretaing a Frame which can expand according to the size of the window

lb0 =Label(f_login,text="Enter Details",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(columnspan=3,row=0,pady="15")
lb1 =Label(f_login,text="Enter ID: ",font="lucida 10 bold").grid(column=0,row=2,pady="4")

lid =StringVar()
e1 =Entry(f_login,textvariable=lid,width="28").grid(column=1,row=2)
lb2 =Label(f_login,text="Enter Password: ",font="lucida 10 bold").grid(column=0,row=3,pady="4")

lpass=StringVar()
e2=Entry(f_login,show="*",textvariable=lpass,width="28").grid(column=1,row=3)
btn=Button(f_login,text="login",bg="green",fg="white",width="10",font="lucida 10 bold",command=Login)
btn.grid(columnspan=3,row=5,pady="10")

f_login.pack(pady="165")

l_login.pack(ipadx="100",fill=BOTH)

''' ************************************************************* Code for Login window Ends ********************************************************'''
''' ************************************************************* Code for Menu Widgit starts ********************************************************'''



def attendance(y):
    l.pack_forget()

    l1.pack(ipadx="150", fill=BOTH)
    if (y == 1):
        fa.pack_forget()
        fd.pack(pady="120")
    if (y == 2):
        fd.pack_forget()
        fa.pack(pady="135")






def Proceed_menu():
    global login_by
    l_login.pack_forget()



    mainmenu = Menu(root)
    root.config(menu=mainmenu)


    m2 = Menu(mainmenu, tearoff=0)
    m2.add_command(label="Detect", command=lambda: attendance(1))
    m2.add_separator()
    m2.add_command(label="View Excel", command=lambda: attendance(2))
    root.config(menu=mainmenu)
    mainmenu.add_cascade(label="Attendance", menu=m2)




''' ************************************************************* Code for Menu Widgit Ends ********************************************************'''

''' ************************************************************* Code for Recording attendance starts ********************************************************'''


def detect1():
    insertdate(desem.get()[0], desection.get())
    dt.detect(desem.get()[0], desection.get())

l = Label(image=photo)
l.pack(ipadx="100", fill=BOTH)

l1 = Label(image=photo)

fd = Frame(l1, pady="25", padx="25")
ld = Label(fd, text="This is Detect Section", bg="orange", fg="blue", font="lucida 10 bold", width="35", pady="4").grid(
    columnspan=3, row=0, pady="15")
l4 = Label(fd, text="Sem", font="lucida 10 bold").grid(column=0, row=1, pady="4")

desem = StringVar()
desem.set("1st sem")  # default value

w1 = OptionMenu(fd, desem, "1st sem", "2nd sem", "3rd sem", "4th sem", "5th sem", "6th sem", "7th sem", "8th sem").grid(
    column=1, row=1, pady="4")
l5 = Label(fd, text="Section", font="lucida 10 bold").grid(column=0, row=2, pady="4")

desection = StringVar()
w2 = OptionMenu(fd, desection, "1", "2").grid(column=1, row=2, pady="4")

b1 = Button(fd, text="Detect", bg="green", fg="white", width="10", font="lucida 10 bold", command=detect1)
b1.grid(columnspan=3, row=3, pady="20")

fd.pack(pady="120")


def insertdate(sem, sec):
    if sem == '1' or sem == '2':
        year = 'first_year'
    elif sem == '3' or sem == '4':
        year = 'second_year'
    elif sem == '5' or sem == '6':
        year = 'third_year'
    else:
        year = 'fourth_year'

    flag = 0
    print('Checking if the date is working or not..')
    for i in D.filterdates():
        if str(i.day) == str(datetime.datetime.today().day) and str(i.month) == str(
                datetime.datetime.today().month) and str(i.year) == str(datetime.datetime.today().year):
            flag = 1
    if flag == 0:
        value = tmsg.askquestion("this is a holiday.... want to continue.")
        if value == "yes":
            wb = load_workbook(f'data/Attendance_xlsx/{year}_{sem}sem_IT{sec}.xlsx')
            if flag == 0:
                print('Date:', str(i)[:11], ' is written in excel and is a holiday')
            else:
                print('Date:', str(i)[:11], ' is written in excel and is a working day')

            sheet = wb.active
            current_row = sheet.max_row
            current_column = sheet.max_column
            print(current_column)
            sheet.column_dimensions['A'].width = 20
            sheet.column_dimensions['B'].width = 20
            sheet.cell(row=1, column=1).value = "Name"
            sheet.cell(row=1, column=2).value = "Enrollment"

            current_row = sheet.max_row
            current_column = sheet.max_column
            # sheet.cell(row=1,column=current_column).width = 20
            sheet.cell(row=1, column=current_column + 1).value = "".join(str(datetime.datetime.today())[:11])

            # save the file
            wb.save('data/Attendance_xlsx/third_year_5sem_IT2.xlsx')


''' ************************************************************* Code for Recording attendance Ends ********************************************************'''

''' ************************************************************* Code for viewing Excel starts ********************************************************'''


def open_excel():
    sem = seme.get()[0]
    sec = sect.get()
    if sem == '1' or sem == '2':
        year = 'first_year'
    elif sem == '3' or sem == '4':
        year = 'second_year'
    elif sem == '5' or sem == '6':
        year = 'third_year'
    else:
        year = 'fourth_year'

    file = "C:\\Users\\HOME\\PycharmProjects\\Face\\data\\Attendance_xlsx\\" + year + "_" + sem + "sem_IT" + sec + ".xlsx"
    os.startfile(file)


fa = Frame(l1, pady="8", padx="20", height=200)
Label(fa, text="Select Year", bg="orange", fg="blue", font="lucida 10 bold", width="30").grid(columnspan=3, row=0,
                                                                                              pady="10")
year = StringVar()
year.set("first_year")
w1 = OptionMenu(fa, year, "first_year", "second_year", "third_year", "fourth_year").grid(columnspan=3, row=1, pady="4")

Label(fa, text="Select Semester", bg="orange", fg="blue", font="lucida 10 bold", width="30").grid(columnspan=3, row=2,
                                                                                                  pady="10")
seme = StringVar()
seme.set("1sem")  # default value
w1 = OptionMenu(fa, seme, "1sem", "2sem", "3sem", "4sem", "5sem", "6sem", "7sem", "8sem").grid(columnspan=3, row=3,
                                                                                               pady="4")

sect = StringVar()
sect.set("nosec")
Label(fa, text="Select Section", bg="orange", fg="blue", font="lucida 10 bold", width="30").grid(columnspan=3, row=5,
                                                                                                 pady="10")
radio = Radiobutton(fa, text="IT-1", variable=sect, value="1").grid(column=0, row=6, pady="4")
radio = Radiobutton(fa, text="IT-2", padx=14, variable=sect, value="2").grid(column=1, row=6, pady="4")
btn = Button(fa, text="show", bg="green", fg="white", width="10", font="lucida 10 bold", command=open_excel)
btn.grid(columnspan=3, row=7, pady="0")
fa.pack(pady="135")
l1.pack(ipadx="150", fill=BOTH)


''' ************************************************************* Code for viewing Excel Ends ********************************************************'''


root.mainloop()
